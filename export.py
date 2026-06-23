"""
app/export.py
─────────────────────────────────────────────────────────────────────────────
Excel-Export für das Camp-Kiosk-Backend (FastAPI + SQLAlchemy + SQLite).

Erzeugt EINE Arbeitsmappe mit drei Tabellenblättern – exakt wie im Frontend:
  1. "Kinder"        – Übersicht Kinder + Guthaben
  2. "Mitarbeiter"   – Übersicht Mitarbeiter + Guthaben (Team kauft auch ein)
  3. "Bestellungen"  – alle Bestellungen mit Positionen

Einbindung in app/main.py:

    from .export import router as export_router
    app.include_router(export_router)

Abhängigkeit:  pip install openpyxl
"""

from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from .database import get_db
from .auth import require_admin          # nur Admins dürfen exportieren
from . import models

router = APIRouter(prefix="/api", tags=["export"])

# ── einheitliche Optik ──────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="E8734A")          # Akzent-Orange
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
CENTER = Alignment(horizontal="center", vertical="center")
THIN = Side(style="thin", color="E8E2D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
EURO_FMT = '#,##0.00 "€"'


def _write_sheet(wb: Workbook, titel: str, kopf: list[str], zeilen: list[list],
                 euro_spalten: set[int] = frozenset()):
    """Legt ein formatiertes Tabellenblatt an (Kopfzeile, Rahmen, €-Format,
    Auto-Breite, eingefrorene Kopfzeile)."""
    ws = wb.create_sheet(title=titel)
    ws.append(kopf)
    for c in range(1, len(kopf) + 1):
        z = ws.cell(row=1, column=c)
        z.fill, z.font, z.alignment, z.border = HEADER_FILL, HEADER_FONT, CENTER, BORDER

    for zeile in zeilen:
        ws.append(zeile)
        r = ws.max_row
        for c in range(1, len(kopf) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if c in euro_spalten:
                cell.number_format = EURO_FMT

    # Spaltenbreiten automatisch an Inhalt anpassen
    for c in range(1, len(kopf) + 1):
        breite = max(
            [len(str(kopf[c - 1]))] +
            [len(str(z[c - 1])) for z in zeilen if c - 1 < len(z)]
        ) + 3
        ws.column_dimensions[get_column_letter(c)].width = min(breite, 55)

    ws.freeze_panes = "A2"
    return ws


@router.get("/export/excel")
def export_excel(db: Session = Depends(get_db), _admin=Depends(require_admin)):
    """Lädt die komplette Abrechnung als .xlsx mit drei Tabellenblättern."""

    wb = Workbook()
    wb.remove(wb.active)  # leeres Default-Blatt entfernen

    # Hilfsfunktion: aktive Ausgaben eines Käufers aufsummieren
    def ausgaben(typ: str, kid: int) -> float:
        s = (
            db.query(models.Bestellung)
            .filter(
                models.Bestellung.kaeufer_typ == typ,
                models.Bestellung.kaeufer_id == kid,
                models.Bestellung.status == "aktiv",
            )
        )
        return round(sum(b.summe for b in s), 2)

    # ── Blatt 1: Kinder ─────────────────────────────────────────────────────
    kinder = db.query(models.Kind).order_by(models.Kind.name).all()
    kinder_rows = [
        [k.name,
         round(k.start_guthaben, 2),
         ausgaben("kind", k.id),
         round(k.guthaben, 2)]
        for k in kinder
    ]
    _write_sheet(
        wb, "Kinder",
        ["Name", "Start-Guthaben", "Ausgaben", "Aktuelles Guthaben"],
        kinder_rows, euro_spalten={2, 3, 4},
    )

    # ── Blatt 2: Mitarbeiter ────────────────────────────────────────────────
    team = db.query(models.Mitarbeiter).order_by(models.Mitarbeiter.name).all()
    team_rows = [
        [m.name, m.rolle,
         round(m.start_guthaben, 2),
         ausgaben("mitarbeiter", m.id),
         round(m.guthaben, 2)]
        for m in team
    ]
    _write_sheet(
        wb, "Mitarbeiter",
        ["Name", "Rolle", "Start-Guthaben", "Ausgaben", "Aktuelles Guthaben"],
        team_rows, euro_spalten={3, 4, 5},
    )

    # ── Blatt 3: Bestellungen ───────────────────────────────────────────────
    bestellungen = (
        db.query(models.Bestellung)
        .order_by(models.Bestellung.zeitstempel)
        .all()
    )
    best_rows = []
    for b in bestellungen:
        artikel = ", ".join(f"{p.menge}× {p.produkt_name}" for p in b.positionen)
        best_rows.append([
            b.zeitstempel.strftime("%d.%m.%Y"),
            b.zeitstempel.strftime("%H:%M"),
            "Kind" if b.kaeufer_typ == "kind" else "Mitarbeiter",
            b.kaeufer_name,
            artikel,
            round(b.summe, 2),
            b.status,
            b.verkauft_von or "",
        ])
    _write_sheet(
        wb, "Bestellungen",
        ["Datum", "Uhrzeit", "Käufer-Typ", "Käufer", "Artikel",
         "Summe", "Status", "Verkauft von"],
        best_rows, euro_spalten={6},
    )

    # ── als Download zurückgeben ────────────────────────────────────────────
    puffer = BytesIO()
    wb.save(puffer)
    puffer.seek(0)
    dateiname = f"camp-kiosk-export_{datetime.now():%Y-%m-%d}.xlsx"
    return StreamingResponse(
        puffer,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f'attachment; filename="{dateiname}"'},
    )
